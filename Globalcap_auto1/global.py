import io
import os
import sys
import webbrowser
from datetime import timedelta
from threading import Timer
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd

# ==============================================================================
# CONFIGURACIÓN DE RUTAS DE PLANTILLAS PARA PYINSTALLER (.EXE)
# ==============================================================================
if getattr(sys, 'frozen', False):
    # Cuando corre como .exe, busca la carpeta 'templates' en la carpeta temporal desempacada
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=template_folder)
else:
    # Cuando corre normal en VS Code / desarrollo
    app = Flask(__name__)

CORS(app)


def open_browser():
    webbrowser.open_new('http://127.0.0.1:5001/')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/unificar', methods=['POST', 'OPTIONS'])
def unificar_archivos():
    if request.method == 'OPTIONS':
        return '', 200

    fecha_inicio_env = request.form.get('fecha_inicio')
    fecha_fin_env = request.form.get('fecha_fin')

    if not fecha_inicio_env or not fecha_fin_env:
        return (
            jsonify({'error': 'Por favor, selecciona un rango de fechas válido.'}),
            400,
        )

    meses_en = {
        1: 'Jan',
        2: 'Feb',
        3: 'Mar',
        4: 'Apr',
        5: 'May',
        6: 'Jun',
        7: 'Jul',
        8: 'Aug',
        9: 'Sep',
        10: 'Oct',
        11: 'Nov',
        12: 'Dec',
    }

    try:
        f_ini_dt = pd.to_datetime(fecha_inicio_env)
        f_fin_dt = pd.to_datetime(fecha_fin_env)

        # AJUSTE 1: Si la fecha seleccionada no es domingo (ej: Lunes), retroceder al domingo anterior
        offset_inicio = (f_ini_dt.dayofweek + 1) % 7
        f_ini_dt_ajustada = f_ini_dt - pd.Timedelta(days=offset_inicio)

        rango_fechas = pd.date_range(
            start=f_ini_dt_ajustada, end=f_fin_dt, freq='W-SUN'
        )
        lista_aux = list(rango_fechas)

        if f_fin_dt.dayofweek == 6 and f_fin_dt not in lista_aux:
            lista_aux.append(f_fin_dt)

        fechas_dt = sorted(list(set(lista_aux)))

        fechas_exigidas_formateadas = []
        mapeo_domingos = {}

        for d in fechas_dt:
            dia_str = str(d.day).zfill(2)
            mes_str = meses_en[d.month]
            anio_str = d.strftime('%y')
            f_formato = f'{dia_str}-{mes_str}-{anio_str}'

            fechas_exigidas_formateadas.append(f_formato)
            mapeo_domingos[d.date()] = f_formato

        anio_referencia = f_ini_dt.year

    except Exception as e:
        return (
            jsonify({'error': f'Error al procesar el rango de fechas: {str(e)}'}),
            400,
        )

    dataframes_listos = []
    universo_programas = {}

    archivos_subidos = request.files.getlist('archivos')
    if not archivos_subidos or archivos_subidos[0].filename == '':
        return (
            jsonify({'error': 'No se recibieron archivos válidos para unificar.'}),
            400,
        )

    def limpiar_valor_metrica(v):
        if pd.isna(v):
            return None
        v_str = str(v).strip().lower()

        if v_str in ['closed', 'closd', 'cloed', 'clsoed', 'cloeded', 'clod']:
            return 'Closed'

        if v_str in ['', 'nan', 'none']:
            return None

        try:
            val_num = pd.to_numeric(v)
            return val_num
        except:
            return str(v).strip()

    def encontrar_fecha_unificada(fecha_raw):
        try:
            dt_raw = None

            if isinstance(fecha_raw, (pd.Timestamp, pd.DatetimeIndex)):
                dt_raw = fecha_raw
            else:
                f_str = str(fecha_raw).split(' ')[0].strip()
                partes = f_str.replace('/', '-').split('-')

                if len(partes) == 2:
                    m, d = int(partes[0]), int(partes[1])
                    dt_raw = pd.Timestamp(year=anio_referencia, month=m, day=d)
                elif len(partes) == 3:
                    dt_raw = pd.to_datetime(f_str, dayfirst=False, errors='coerce')

            if pd.isna(dt_raw) or dt_raw is None:
                return None

            if dt_raw.dayofweek == 0:
                dt_raw = dt_raw - pd.Timedelta(days=1)

            min_diff = 999
            fecha_cercana = None

            for f_ex in fechas_dt:
                diff = abs((dt_raw.date() - f_ex.date()).days)
                if diff <= 3 and diff < min_diff:
                    min_diff = diff
                    fecha_cercana = mapeo_domingos[f_ex.date()]

            return fecha_cercana
        except:
            return None

    # PROCESAMIENTO DE ARCHIVOS EXCEL
    for file in archivos_subidos:
        try:
            file_bytes = file.read()
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
            pestaña_objetivo = excel_file.sheet_names[0]

            for sheet in excel_file.sheet_names:
                if sheet.strip() in [
                    'Reporte',
                    'Resumen Continuous 2026',
                    'Resumen Continuo 2026',
                    'Sheet1',
                    'Hoja1',
                    'Program H2R',
                ]:
                    pestaña_objetivo = sheet
                    break

            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=pestaña_objetivo)
            df.columns = [str(c).replace('\xa0', ' ').strip() for c in df.columns]

            col_country = next(
                (c for c in df.columns if c.lower() == 'country'), None
            )
            col_client = next(
                (
                    c
                    for c in df.columns
                    if c.lower() in ['client', 'client name', 'cliente']
                ),
                None,
            )
            col_prog_h2r = next(
                (
                    c
                    for c in df.columns
                    if c.lower()
                    in [
                        'program h2r',
                        'program_h2r',
                        'program',
                        'h2r department',
                        'programa',
                    ]
                ),
                None,
            )
            col_fecha_origen = next(
                (
                    c
                    for c in df.columns
                    if c.lower() in ['fecha', 'description', 'metric']
                ),
                None,
            )

            col_b = next(
                (c for c in df.columns if 'budget' in c.lower()), 'Budget'
            )
            col_f = next(
                (c for c in df.columns if 'forecast' in c.lower()), 'Forecast'
            )
            col_r = next(
                (c for c in df.columns if 'required' in c.lower()), 'Required'
            )
            col_p = next(
                (c for c in df.columns if c.lower() in ['projected', 'project']),
                'Projected',
            )

            columnas_fechas = [
                c
                for c in df.columns
                if c
                not in [
                    col_country,
                    col_client,
                    col_prog_h2r,
                    col_fecha_origen,
                    'Program ID (Client Name)',
                    'Description',
                ]
                and any(char.isdigit() for char in str(c))
            ]

            es_horizontal = len(columnas_fechas) > 0

            # CASO A: ESTRUCTURA HORIZONTAL
            if es_horizontal:
                for _, fila in df.iterrows():
                    client_val = (
                        str(fila.get(col_client, '')).replace('\xa0', ' ').strip()
                        if col_client
                        else ''
                    )
                    progh2r_val = (
                        str(fila.get(col_prog_h2r, '')).replace('\xa0', ' ').strip()
                        if col_prog_h2r
                        else ''
                    )
                    country_val = (
                        str(fila.get(col_country, '')).replace('\xa0', ' ').strip()
                        if col_country
                        else ''
                    )

                    if client_val.lower() in ['nan', 'none']:
                        client_val = ''
                    if progh2r_val.lower() in ['nan', 'none']:
                        progh2r_val = ''
                    if country_val.lower() in ['nan', 'none']:
                        country_val = ''

                    llave = (country_val, client_val, progh2r_val)
                    universo_programas[llave] = True

                    tipo_metrica_fila = (
                        str(fila.get(col_fecha_origen, ''))
                        .replace('\xa0', ' ')
                        .strip()
                        .lower()
                    )

                    for col_f_matriz in columnas_fechas:
                        f_unificada = encontrar_fecha_unificada(col_f_matriz)
                        if not f_unificada:
                            continue

                        val_celda = limpiar_valor_metrica(fila.get(col_f_matriz))

                        b_v, f_v, r_v, p_v = None, None, None, None

                        if 'budg' in tipo_metrica_fila:
                            b_v = val_celda
                        elif 'fore' in tipo_metrica_fila:
                            f_v = val_celda
                        elif 'req' in tipo_metrica_fila:
                            r_v = val_celda
                        elif 'proj' in tipo_metrica_fila:
                            p_v = val_celda

                        if val_celda is not None:
                            dataframes_listos.append({
                                'Country': country_val,
                                'Client': client_val,
                                'Program H2R': progh2r_val,
                                'Weeks': f_unificada,
                                'Budget': b_v,
                                'Forecast': f_v,
                                'Required': r_v,
                                'Projected': p_v,
                            })

            # CASO B: ESTRUCTURA VERTICAL ESTÁNDAR
            elif col_fecha_origen:
                for _, fila in df.iterrows():
                    client_val = (
                        str(fila.get(col_client, '')).replace('\xa0', ' ').strip()
                        if col_client
                        else ''
                    )
                    progh2r_val = (
                        str(fila.get(col_prog_h2r, '')).replace('\xa0', ' ').strip()
                        if col_prog_h2r
                        else ''
                    )

                    if client_val.lower() in ['nan', 'none']:
                        client_val = ''
                    if progh2r_val.lower() in ['nan', 'none']:
                        progh2r_val = ''

                    country_val = (
                        str(fila.get(col_country, '')).replace('\xa0', ' ').strip()
                        if col_country
                        else ''
                    )
                    if country_val.lower() in ['nan', 'none']:
                        country_val = ''

                    llave = (country_val, client_val, progh2r_val)
                    universo_programas[llave] = True

                    f_raw = (
                        str(fila.get(col_fecha_origen, '')).replace('\xa0', ' ').strip()
                    )
                    f_unificada = encontrar_fecha_unificada(f_raw)
                    if not f_unificada:
                        continue

                    dataframes_listos.append({
                        'Country': country_val,
                        'Client': client_val,
                        'Program H2R': progh2r_val,
                        'Weeks': f_unificada,
                        'Budget': limpiar_valor_metrica(fila.get(col_b)),
                        'Forecast': limpiar_valor_metrica(fila.get(col_f)),
                        'Required': limpiar_valor_metrica(fila.get(col_r)),
                        'Projected': limpiar_valor_metrica(fila.get(col_p)),
                    })

        except Exception as e:
            print(f'[Error] Archivo saltado: {str(e)}', file=sys.stderr, flush=True)

    # UNIFICACIÓN Y MATRIZ FINAL
    if universo_programas:
        registros_molde = []
        for c_val, cl_val, ph2r_val in universo_programas.keys():
            for f_ex in fechas_exigidas_formateadas:
                registros_molde.append({
                    'Country': c_val,
                    'Client': cl_val,
                    'Program H2R': ph2r_val,
                    'Weeks': f_ex,
                })
        df_maestro = pd.DataFrame(registros_molde)

        if dataframes_listos:
            df_extraido = pd.DataFrame(dataframes_listos)

            def combiner_valores(series):
                validos = [
                    v
                    for v in series.dropna()
                    if str(v).strip().lower() not in ['none', 'nan', '']
                ]
                if validos:
                    return validos[0]
                return None

            df_extraido = df_extraido.groupby(
                ['Country', 'Client', 'Program H2R', 'Weeks'],
                as_index=False,
                dropna=False,
            ).agg({
                'Budget': combiner_valores,
                'Forecast': combiner_valores,
                'Required': combiner_valores,
                'Projected': combiner_valores,
            })

            df_resultado = pd.merge(
                df_maestro,
                df_extraido,
                on=['Country', 'Client', 'Program H2R', 'Weeks'],
                how='left',
            )
        else:
            df_resultado = df_maestro
            for m_col in ['Budget', 'Forecast', 'Required', 'Projected']:
                df_resultado[m_col] = None

        df_resultado['fecha_aux_sort'] = pd.to_datetime(
            df_resultado['Weeks'], format='%d-%b-%y', errors='coerce'
        )
        df_resultado = df_resultado.sort_values(
            by=['Country', 'Program H2R', 'Client', 'fecha_aux_sort']
        ).reset_index(drop=True)
        df_resultado.drop(columns=['fecha_aux_sort'], inplace=True)

        for col in ['Budget', 'Forecast', 'Required', 'Projected']:
            df_resultado[col] = df_resultado[col].apply(
                lambda x: (
                    ''
                    if pd.isna(x) or str(x).strip().lower() in ['nan', 'none', '']
                    else (
                        'Closed'
                        if str(x).strip().lower() == 'closed'
                        else int(round(pd.to_numeric(x, errors='coerce')))
                    )
                )
            )

        columnas_finales = [
            'Country',
            'Program H2R',
            'Client',
            'Weeks',
            'Budget',
            'Forecast',
            'Required',
            'Projected',
        ]
        df_resultado = df_resultado[columnas_finales]

        salida_bytes = io.BytesIO()
        with pd.ExcelWriter(salida_bytes, engine='openpyxl') as writer:
            df_resultado.to_excel(
                writer, sheet_name='Reporte Consolidado', index=False
            )
            worksheet = writer.sheets['Reporte Consolidado']

            font_headers = Font(name='Segoe UI', size=11, bold=True)
            font_rows = Font(name='Segoe UI', size=11, bold=False)
            font_weeks_col = Font(
                name='Segoe UI', size=11, bold=False, color='FFFFFF'
            )

            fill_weeks = PatternFill(
                start_color='1F4E5F', end_color='1F4E5F', fill_type='solid'
            )

            borde_fino = Border(
                left=Side(style='medium', color='7F7F7F'),
                right=Side(style='medium', color='7F7F7F'),
                top=Side(style='medium', color='7F7F7F'),
                bottom=Side(style='medium', color='7F7F7F'),
            )

            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')

            for cell in worksheet[1]:
                cell.font = font_headers
                cell.border = borde_fino
                cell.alignment = align_center

            for row in range(2, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = borde_fino

                    if col in [1, 2, 3]:
                        cell.font = font_rows
                        cell.alignment = align_left
                    elif col == 4:
                        cell.font = font_weeks_col
                        cell.fill = fill_weeks
                        cell.alignment = align_center
                    else:
                        if cell.value not in [None, '', 'Closed']:
                            try:
                                cell.value = int(cell.value)
                                cell.number_format = '#,##0'
                            except:
                                pass
                        cell.font = font_rows
                        cell.alignment = align_center

            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        salida_bytes.seek(0)

        return send_file(
            salida_bytes,
            mimetype=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
            as_attachment=True,
            download_name='GlobalCap.xlsx',
        )
    else:
        return 'No se encontraron datos procesables.', 500


if __name__ == '__main__':
    # Dispara la apertura del navegador a los 1.5 segundos
    Timer(1.5, open_browser).start()
    # Ejecuta Flask con debug deshabilitado para que no interfiera con PyInstaller
    app.run(host='127.0.0.1', port=5001, debug=False)
